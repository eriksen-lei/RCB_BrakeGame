import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "预算清单"

# Colors
COLOR_TITLE_BG   = "1F3864"
COLOR_SECTION_BG = "2E75B6"
COLOR_HEADER_BG  = "BDD7EE"
COLOR_SUBTOTAL_BG= "D6E4F0"
COLOR_TOTAL_BG   = "FFE699"
COLOR_NOTE_BG    = "F2F2F2"
COLOR_WHITE      = "FFFFFF"
COLOR_DARK       = "1F1F1F"

thin = Side(style="thin", color="AAAAAA")
medium = Side(style="medium", color="888888")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)

def cell_style(cell, bold=False, size=11, color=COLOR_DARK, bg=None,
               align="left", valign="center", border=True, wrap=False):
    cell.font = Font(name="微软雅黑", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if border:
        cell.border = thin_border

def write_row(ws, row, values, bold=False, size=11, color=COLOR_DARK, bg=None,
              align_list=None, border=True, wrap=False):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i+1, value=v)
        al = align_list[i] if align_list else "left"
        cell_style(c, bold=bold, size=size, color=color, bg=bg,
                   align=al, border=border, wrap=wrap)

# Column widths
col_widths = [28, 38, 8, 8, 8]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 20

# ── Title ──────────────────────────────────────────────────────
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "摩托车制动体感系统 · 完整预算清单（¥4,000）"
cell_style(c, bold=True, size=16, color=COLOR_WHITE, bg=COLOR_TITLE_BG, align="center")

ws.merge_cells("A2:E2")
c = ws["A2"]
c.value = "传感器方案：微型霍尔传感器安装于卡钳内（主）+ 液压压力传感器（备）"
cell_style(c, bold=False, size=10, color=COLOR_WHITE, bg=COLOR_TITLE_BG, align="center")

# ── Section 1 ──────────────────────────────────────────────────
r = 3
ws.merge_cells(f"A{r}:E{r}")
c = ws[f"A{r}"]
c.value = "一、硬件物料"
cell_style(c, bold=True, size=12, color=COLOR_WHITE, bg=COLOR_SECTION_BG, align="left")
ws.row_dimensions[r].height = 22

r += 1
write_row(ws, r, ["零件", "规格/型号", "数量", "单价", "小计"],
          bold=True, bg=COLOR_HEADER_BG,
          align_list=["left","left","center","center","center"])
ws.row_dimensions[r].height = 18

hardware = [
    ("线性霍尔传感器", "AH3505/SS495A，固定于卡钳内测量活塞推出距离，1主4备", 5, 5, 25),
    ("钕磁铁 N52薄片", "8×3×2mm，粘贴于卡钳活塞端面，随活塞移动", 20, 2, 40),
    ("液压压力传感器", "0–1MPa，G1/8，0–5V，备用方案", 1, 60, 60),
    ("T型三通接头", "G1/8，不锈钢，备用方案油路接入", 2, 20, 40),
    ("传感器防护套", "硅胶防油污套，卡钳内环境保护", 5, 8, 40),
    ("Arduino Pro Micro", "5V/16MHz 兼容版，1主1备", 2, 35, 70),
    ("PCB打板", "信号调理+运放放大+滤波+供电整合，嘉立创5片", 1, 120, 120),
    ("3D打印支架", "PETG耐油耐热，卡钳内传感器固定座，4次迭代", 4, 55, 220),
    ("防水接线端子", "IP67，适配卡钳附近潮湿/油污环境", 1, 25, 25),
    ("屏蔽信号线", "双绞屏蔽线，抗刹车系统电磁干扰", 3, 10, 30),
    ("USB数据线", "Type-C，1.5m", 2, 10, 20),
    ("杜邦线+面包板套装", "原型调试用", 1, 20, 20),
    ("扎带+螺丝+热缩管", "安装固定耗材包", 1, 25, 25),
    ("电子散件包", "电阻/电容/LDO稳压/运放IC，信号调理用", 1, 40, 40),
    ("环氧树脂胶（AB胶）", "传感器+磁铁卡钳内永久固定", 2, 8, 16),
    ("耐油密封胶", "卡钳开孔走线密封用", 1, 19, 19),
]

for item in hardware:
    r += 1
    write_row(ws, r, list(item),
              align_list=["left","left","center","center","center"])
    ws.row_dimensions[r].height = 16

r += 1
ws.merge_cells(f"A{r}:D{r}")
c = ws[f"A{r}"]
c.value = "硬件合计"
cell_style(c, bold=True, bg=COLOR_SUBTOTAL_BG, align="right")
ws.cell(row=r, column=5, value=1010).font = Font(name="微软雅黑", bold=True)
ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=COLOR_SUBTOTAL_BG)
ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
ws.cell(row=r, column=5).border = thin_border
ws.row_dimensions[r].height = 18

# ── Section 2 ──────────────────────────────────────────────────
r += 1
ws.merge_cells(f"A{r}:E{r}")
c = ws[f"A{r}"]
c.value = "二、人工与服务费"
cell_style(c, bold=True, size=12, color=COLOR_WHITE, bg=COLOR_SECTION_BG, align="left")
ws.row_dimensions[r].height = 22

r += 1
write_row(ws, r, ["项目", "工作内容", "工时", "单价", "金额"],
          bold=True, bg=COLOR_HEADER_BG,
          align_list=["left","left","center","center","center"])

labor = [
    ("硬件方案设计", "卡钳内布局分析、磁路间距计算、方案文档", "1天", "", 350),
    ("PCB设计", "原理图+运放放大电路+Layout+嘉立创下单全流程", "1天", "", 350),
    ("3D支架建模", "卡钳内腔尺寸测绘+固定座建模+4次迭代沟通", "1.5天", "", 350),
    ("硬件装配调试", "传感器卡钳内安装、磁铁定位、防水接线验证", "1天", "", 350),
    ("信号校准", "霍尔线性区标定、运放增益调整、ADC映射", "0.5天", "", 180),
    ("Arduino固件开发", "采样/滑动滤波/USB HID游戏手柄模拟", "1天", "", 300),
    ("Python通信模块", "HID读取、实时数据流处理", "0.5天", "", 150),
    ("游戏逻辑开发", "状态机、反应时间计算、评分算法", "1.5天", "", 350),
    ("UI界面开发", "力度条、倒计时动画、结果页、排行榜", "1天", "", 280),
    ("数据存储与导出", "JSON本地排行榜、CSV数据导出", "0.5天", "", 120),
    ("软硬件联调", "全链路联调、体验优化、延迟测试", "1天", "", 250),
    ("项目管理+交付文档", "需求沟通、安装手册、验收报告", "全程", "", 250),
]

for item in labor:
    r += 1
    write_row(ws, r, list(item),
              align_list=["left","left","center","center","center"])
    ws.row_dimensions[r].height = 16

r += 1
ws.merge_cells(f"A{r}:D{r}")
c = ws[f"A{r}"]
c.value = "人工合计"
cell_style(c, bold=True, bg=COLOR_SUBTOTAL_BG, align="right")
ws.cell(row=r, column=5, value=2730).font = Font(name="微软雅黑", bold=True)
ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=COLOR_SUBTOTAL_BG)
ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
ws.cell(row=r, column=5).border = thin_border
ws.row_dimensions[r].height = 18

# ── Section 3 ──────────────────────────────────────────────────
r += 1
ws.merge_cells(f"A{r}:E{r}")
c = ws[f"A{r}"]
c.value = "三、风险预留"
cell_style(c, bold=True, size=12, color=COLOR_WHITE, bg=COLOR_SECTION_BG, align="left")
ws.row_dimensions[r].height = 22

r += 1
write_row(ws, r, ["项目", "说明", "", "", "金额"],
          bold=True, bg=COLOR_HEADER_BG,
          align_list=["left","left","left","left","center"])

risks = [
    ("硬件返工预留", "PCB二次打样或磁路重新设计", 100),
    ("支架超额迭代", "卡钳内腔尺寸差异导致的额外打印", 80),
    ("密封/防油返工", "卡钳环境恶劣，走线密封可能需重做", 50),
    ("现场支持/展会部署", "活动现场调试支援（半天）", 30),
]

for item in risks:
    r += 1
    ws.merge_cells(f"B{r}:D{r}")
    ws.cell(row=r, column=1, value=item[0])
    cell_style(ws.cell(row=r, column=1), align="left")
    ws.cell(row=r, column=2, value=item[1])
    cell_style(ws.cell(row=r, column=2), align="left")
    for col in [3, 4]:
        cell_style(ws.cell(row=r, column=col), align="left")
    ws.cell(row=r, column=5, value=item[2])
    cell_style(ws.cell(row=r, column=5), align="center")
    ws.row_dimensions[r].height = 16

r += 1
ws.merge_cells(f"A{r}:D{r}")
c = ws[f"A{r}"]
c.value = "预留合计"
cell_style(c, bold=True, bg=COLOR_SUBTOTAL_BG, align="right")
ws.cell(row=r, column=5, value=260).font = Font(name="微软雅黑", bold=True)
ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=COLOR_SUBTOTAL_BG)
ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
ws.cell(row=r, column=5).border = thin_border
ws.row_dimensions[r].height = 18

# ── Total ──────────────────────────────────────────────────────
r += 1
ws.merge_cells(f"A{r}:E{r}")
ws.row_dimensions[r].height = 6  # spacer

r += 1
ws.merge_cells(f"A{r}:D{r}")
c = ws[f"A{r}"]
c.value = "总计"
cell_style(c, bold=True, size=13, color=COLOR_DARK, bg=COLOR_TOTAL_BG, align="right")
ws.cell(row=r, column=5, value=4000).font = Font(name="微软雅黑", bold=True, size=13)
ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
ws.cell(row=r, column=5).border = medium_border
ws.row_dimensions[r].height = 24

# ── Notes ──────────────────────────────────────────────────────
r += 2
notes = [
    "备注：",
    "1. 传感器安装于卡钳内部，霍尔传感器固定于卡钳活塞侧壁，磁铁粘贴于活塞端面，刹车时活塞推出，磁场变化映射为制动位移量。",
    "2. 卡钳内环境恶劣（制动液/高温/震动），新增防水接线端子（IP67）、屏蔽信号线、耐油密封胶、防护硅胶套。",
    "3. PCB升级加入运放放大电路：卡钳内活塞行程小，霍尔信号幅度偏弱，需放大处理后送ADC。",
    "4. 3D打印改用PETG材质（耐油耐热），迭代4次，卡钳内腔公差紧，首版难以直接装配。",
    "5. Arduino Pro Micro 原生USB HID，直接模拟游戏手柄，省去串口驱动安装问题。",
]
for note in notes:
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = note
    bold = note == "备注："
    cell_style(c, bold=bold, size=9, color="444444", bg=COLOR_NOTE_BG,
               align="left", border=False, wrap=True)
    ws.row_dimensions[r].height = 28 if len(note) > 40 else 16
    r += 1

path = r"E:\RCB_Project_Files\预算清单_4000元.xlsx"
wb.save(path)
print(f"saved: {path}")
